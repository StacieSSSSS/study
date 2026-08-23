"""Read the manually refreshed Wind workbook with conservative availability dates."""

from __future__ import annotations

import hashlib
import os
from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path
from typing import Any, cast

import numpy as np
import pandas as pd
import yaml
from openpyxl import load_workbook
from pandas.tseries.offsets import BDay

RELEASE_DICTIONARY_PATH = Path(__file__).with_name("release_dictionary.yaml")


@dataclass(frozen=True)
class ManualExcelBundle:
    """Cleaned strategy panel plus the audits needed to reproduce it."""

    market_data: pd.DataFrame
    macro_observations: pd.DataFrame
    price_quality: pd.DataFrame
    macro_update_audit: pd.DataFrame
    active_universe: list[str]
    workbook_sha256: str
    workbook_path: Path


def load_release_dictionary(path: Path = RELEASE_DICTIONARY_PATH) -> dict[str, Any]:
    with open(path, encoding="utf-8") as handle:
        return cast(dict[str, Any], yaml.safe_load(handle))


def resolve_workbook_path(config: dict[str, Any], override: str | Path | None = None) -> Path:
    manual = config["data"]["manual_excel"]
    supplied = override or os.getenv(str(manual["path_env"])) or manual["default_path"]
    path = Path(supplied).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(
            f"manual Wind workbook not found at {path}; pass --workbook or set {manual['path_env']}"
        )
    return path


def _sheet_rows(path: Path, sheet_name: str) -> list[tuple[Any, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"required sheet missing: {sheet_name}")
        return list(workbook[sheet_name].iter_rows(values_only=True))
    finally:
        workbook.close()


def _following_weekday(value: pd.Timestamp) -> pd.Timestamp:
    result = cast(pd.Timestamp, value.normalize())
    while result.weekday() >= 5:
        result += pd.Timedelta(days=1)
    return cast(pd.Timestamp, result)


def _next_month_start(reference: pd.Timestamp) -> pd.Timestamp:
    year = reference.year + int(reference.month == 12)
    month = 1 if reference.month == 12 else reference.month + 1
    return cast(pd.Timestamp, pd.Timestamp(year=year, month=month, day=1))


def _rule_date(reference: pd.Timestamp, rule: dict[str, Any]) -> pd.Timestamp:
    kind = str(rule["kind"])
    if kind == "same_month_last_weekday":
        result = cast(pd.Timestamp, reference + pd.offsets.MonthEnd(0))
        while result.weekday() >= 5:
            result -= pd.Timedelta(days=1)
        return cast(pd.Timestamp, result)

    month_start = _next_month_start(reference)
    if kind in {"next_month_day", "quarter_next_month_day"}:
        result = cast(pd.Timestamp, month_start + pd.Timedelta(days=int(rule["day"]) - 1))
        return _following_weekday(result)
    if kind == "next_month_business_ordinal":
        ordinal = int(rule["ordinal"])
        return cast(pd.Timestamp, pd.bdate_range(month_start, periods=ordinal)[-1])
    if kind in {"next_month_last_weekday", "quarter_next_month_last_weekday"}:
        result = cast(pd.Timestamp, month_start + pd.offsets.MonthEnd(0))
        while result.weekday() >= 5:
            result -= pd.Timedelta(days=1)
        return cast(pd.Timestamp, result)
    raise ValueError(f"unsupported conservative release rule: {kind}")


def _release_timestamp(
    reference: pd.Timestamp,
    series: dict[str, Any],
    group: dict[str, Any],
) -> pd.Timestamp:
    latest_reference = pd.Timestamp(series["latest_reference_period"])
    status = str(series["reconciliation_status"])
    if reference == latest_reference and status == "official_match":
        exact = group.get("latest_official_release_ts")
        if exact:
            return cast(pd.Timestamp, pd.Timestamp(exact))

    if reference == latest_reference and status in {
        "date_match_time_unverified",
        "late_vendor_update",
    }:
        update_date = cast(pd.Timestamp, pd.Timestamp(series["latest_workbook_update"]))
        return cast(
            pd.Timestamp,
            pd.Timestamp(
                datetime.combine(update_date.date(), time(23, 59)),
                tz="Asia/Hong_Kong",
            ),
        )

    release_date = _rule_date(reference, cast(dict[str, Any], group["conservative_rule"]))
    local_time = time.fromisoformat(str(group["typical_time"]))
    return cast(
        pd.Timestamp,
        pd.Timestamp(
        datetime.combine(release_date.date(), local_time),
        tz=str(group["timezone"]),
        ),
    )


def _availability_session(timestamp: pd.Timestamp, dictionary: dict[str, Any]) -> pd.Timestamp:
    clock = dictionary["decision_clock"]
    local = timestamp.tz_convert(str(clock["timezone"]))
    cutoff = time.fromisoformat(str(clock["decision_time"]))
    session = local.tz_localize(None).normalize()
    if session.weekday() >= 5 or local.time() > cutoff:
        session = session + BDay(1)
    return cast(pd.Timestamp, _following_weekday(session))


def _read_price_panel(
    rows: list[tuple[Any, ...]], config: dict[str, Any]
) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    manual = config["data"]["manual_excel"]
    headers = list(rows[1])
    raw = pd.DataFrame(rows[3:], columns=headers)
    raw = raw.rename(columns={headers[0]: "date"})
    raw["date"] = pd.to_datetime(raw["date"], errors="coerce")
    raw = raw.dropna(subset=["date"]).set_index("date").sort_index()
    if raw.index.has_duplicates:
        raise ValueError("Price_raw contains duplicate dates")

    selected: dict[str, pd.Series] = {}
    quality_rows: list[dict[str, Any]] = []
    last_dates: list[pd.Timestamp] = []
    active: list[str] = []
    for instrument, source_name in manual["price_map"].items():
        if source_name not in raw.columns:
            quality_rows.append(
                {"instrument": instrument, "source_name": source_name, "status": "missing_column"}
            )
            continue
        values = cast(pd.Series, pd.to_numeric(cast(pd.Series, raw[source_name]), errors="coerce"))
        zero_count = int(values.eq(0).sum())
        values = values.mask(values.eq(0)).loc[lambda item: item.index.weekday < 5]
        nonzero = values.dropna()
        if nonzero.empty:
            quality_rows.append(
                {"instrument": instrument, "source_name": source_name, "status": "no_observations"}
            )
            continue
        selected[instrument] = values
        last_dates.append(cast(pd.Timestamp, nonzero.index[-1]))
        active.append(instrument)
        quality_rows.append(
            {
                "instrument": instrument,
                "source_name": source_name,
                "status": "usable",
                "zero_as_missing": zero_count,
                "first_observation": nonzero.index[0],
                "last_observation": nonzero.index[-1],
                "raw_nonzero_observations": len(nonzero),
            }
        )
    if not selected:
        raise ValueError("none of the configured manual price columns were found")

    common_end = min(last_dates)
    common_start = max(cast(pd.Timestamp, values.dropna().index[0]) for values in selected.values())
    dates = pd.bdate_range(common_start, common_end)
    max_fill = int(manual["max_price_forward_fill_business_days"])
    panel = pd.DataFrame(index=dates)
    panel.index.name = "date"
    for instrument, values in selected.items():
        close = values.reindex(dates).ffill(limit=max_fill)
        panel[f"{instrument}__close"] = close
        spec = cast(dict[str, Any], config["instruments"][instrument])
        return_model = str(spec.get("return_model", "log_price"))
        if return_model == "log_price":
            instrument_return = np.log(close / close.shift(1))
        elif return_model == "yield_duration":
            # Workbook yields are percentage points. Positive strategy weight
            # means long duration / receive fixed, so falling yields earn PnL.
            instrument_return = -float(spec["duration"]) * close.diff() / 100.0
        else:
            raise ValueError(f"unsupported return_model for {instrument}: {return_model}")
        panel[f"{instrument}__return"] = instrument_return
        panel[f"{instrument}__carry"] = np.nan
        after_fill = int(close.isna().sum())
        for row in quality_rows:
            if row.get("instrument") == instrument:
                row["business_missing_after_limited_fill"] = after_fill
                row["common_backtest_end"] = common_end
                row["return_model"] = return_model
                break
    return panel, pd.DataFrame(quality_rows), active


def _period_is_usable(reference: pd.Timestamp, series: dict[str, Any]) -> bool:
    if reference > pd.Timestamp(series["latest_reference_period"]):
        return False
    policy = str(series["period_policy"])
    if policy == "quarter_end_only" and reference.month not in {3, 6, 9, 12}:
        return False
    return not (policy == "jan_feb_combined" and reference.month == 1)


def _transform(values: pd.Series, method: str) -> pd.Series:
    if method == "distance_from_50":
        return values - 50.0
    if method.startswith("change_"):
        return values.diff(int(method.split("_", maxsplit=1)[1]))
    if method in {"level", "winsorized_level"}:
        return values
    raise ValueError(f"unsupported macro transformation: {method}")


def _read_macro_observations(
    rows: list[tuple[Any, ...]], dictionary: dict[str, Any]
) -> tuple[pd.DataFrame, pd.DataFrame]:
    headers = list(rows[1])
    dates = pd.to_datetime([row[0] for row in rows[4:]], errors="coerce")
    observations: list[dict[str, Any]] = []
    audits: list[dict[str, Any]] = []
    groups = dictionary["release_groups"]

    for series in dictionary["series"]:
        column = int(series["workbook_column"])
        actual_name = headers[column - 1]
        if actual_name != series["workbook_name"]:
            raise ValueError(
                f"Macro column {column} changed: expected {series['workbook_name']!r}, got {actual_name!r}"
            )
        group = groups[series["release_group"]]
        official_ts = group.get("latest_official_release_ts")
        official_date = (
            pd.Timestamp(official_ts).date().isoformat()
            if official_ts
            else str(group.get("latest_official_release_date", ""))
        )
        workbook_update = pd.Timestamp(series["latest_workbook_update"])
        official_date_ts = pd.Timestamp(official_date) if official_date else pd.NaT
        audits.append(
            {
                "series_id": series["id"],
                "workbook_column": column,
                "workbook_name": actual_name,
                "latest_reference_period": series["latest_reference_period"],
                "workbook_update_date": workbook_update.date().isoformat(),
                "official_release_date": official_date,
                "workbook_minus_official_days": (
                    int((workbook_update - official_date_ts).days)
                    if not pd.isna(official_date_ts)
                    else np.nan
                ),
                "status": series["reconciliation_status"],
                "note": series.get("reconciliation_note", ""),
                "official_url": group["official_url"],
            }
        )
        if not bool(series.get("include", True)):
            continue

        raw_values = cast(
            pd.Series,
            pd.to_numeric(
                pd.Series([row[column - 1] for row in rows[4:]], index=dates),
                errors="coerce",
            ),
        )
        low, high = (float(value) for value in series["plausible_range"])
        raw_values = raw_values.mask((raw_values < low) | (raw_values > high))
        if not bool(series["zero_is_valid"]):
            raw_values = raw_values.mask(raw_values.eq(0))

        kept: list[dict[str, Any]] = []
        for reference, value in raw_values.items():
            reference_ts = cast(pd.Timestamp, reference)
            if pd.isna(value) or not _period_is_usable(reference_ts, series):
                continue
            release_ts = _release_timestamp(reference_ts, series, group)
            kept.append(
                {
                    "series_id": series["id"],
                    "workbook_name": actual_name,
                    "reference_date": reference_ts,
                    "value": float(value),
                    "release_ts": release_ts,
                    "available_session": _availability_session(release_ts, dictionary),
                    "release_group": series["release_group"],
                    "vintage_status": dictionary["vintage_policy"]["workbook_vintage"],
                    "strict_point_in_time_eligible": False,
                }
            )

        if not kept:
            continue
        event_frame = pd.DataFrame(kept).sort_values("reference_date")
        transformed = _transform(cast(pd.Series, event_frame["value"]), str(series["transformation"]))
        window = 36 if series["frequency"] == "monthly" else 20
        minimum = 12 if series["frequency"] == "monthly" else 6
        mean = transformed.rolling(window, min_periods=minimum).mean()
        std = transformed.rolling(window, min_periods=minimum).std(ddof=1)
        event_frame["transformed_value"] = transformed
        event_frame["feature_z"] = ((transformed - mean) / std).clip(-3.0, 3.0)
        observations.extend(event_frame.to_dict(orient="records"))

    output = pd.DataFrame(observations).sort_values(["available_session", "series_id"])
    audit = pd.DataFrame(audits).sort_values("workbook_column")
    return output.reset_index(drop=True), audit.reset_index(drop=True)


def _attach_macro_baskets(
    panel: pd.DataFrame,
    observations: pd.DataFrame,
    config: dict[str, Any],
    active: list[str],
) -> pd.DataFrame:
    result = panel.copy()
    baskets = config["data"]["manual_excel"]["macro_baskets"]
    minimum = int(config["data"]["manual_excel"]["minimum_macro_components"])
    dates = cast(pd.DatetimeIndex, result.index)
    feature_series: dict[str, pd.Series] = {}
    for series_id, group in observations.dropna(subset=["feature_z"]).groupby("series_id"):
        event = cast(pd.DataFrame, group).sort_values("available_session")
        values = cast(pd.Series, event.groupby("available_session")["feature_z"].last())
        feature_series[str(series_id)] = values.reindex(dates).ffill()

    for instrument in active:
        weights = cast(dict[str, float], baskets.get(instrument, {}))
        numerator = pd.Series(0.0, index=dates)
        denominator = pd.Series(0.0, index=dates)
        count = pd.Series(0, index=dates)
        for series_id, weight in weights.items():
            if series_id not in feature_series:
                continue
            feature = feature_series[series_id]
            available = feature.notna()
            numerator = numerator.add(feature.fillna(0.0) * float(weight), fill_value=0.0)
            denominator = denominator.add(available.astype(float) * abs(float(weight)), fill_value=0.0)
            count = count.add(available.astype(int), fill_value=0)
        macro = cast(
            pd.Series,
            numerator.div(denominator.replace(0.0, np.nan)).where(count >= minimum),
        )
        result[f"{instrument}__macro"] = macro
    return result


def load_manual_excel_bundle(
    config: dict[str, Any],
    workbook_path: str | Path | None = None,
    dictionary_path: Path = RELEASE_DICTIONARY_PATH,
) -> ManualExcelBundle:
    """Load, clean, timestamp, and audit a manual Wind workbook."""
    dictionary = load_release_dictionary(dictionary_path)
    path = resolve_workbook_path(config, workbook_path)
    price_rows = _sheet_rows(path, str(dictionary["workbook"]["price_sheet"]))
    macro_rows = _sheet_rows(path, str(dictionary["workbook"]["macro_sheet"]))
    panel, price_quality, active = _read_price_panel(price_rows, config)
    observations, update_audit = _read_macro_observations(macro_rows, dictionary)
    panel = _attach_macro_baskets(panel, observations, config, active)
    return ManualExcelBundle(
        market_data=panel,
        macro_observations=observations,
        price_quality=price_quality,
        macro_update_audit=update_audit,
        active_universe=active,
        workbook_sha256=hashlib.sha256(path.read_bytes()).hexdigest(),
        workbook_path=path,
    )
