"""Profile and schema-check a manually refreshed Raw_wind.xlsx workbook."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, cast

import pandas as pd
from openpyxl import load_workbook

from strategies.wind_macro_daily.backtest.run import _write_csv
from strategies.wind_macro_daily.data.excel_loader import load_manual_excel_bundle
from strategies.wind_macro_daily.data.loader import load_config


def _sheet_profile(path: Path, sheet_name: str, data_start_row: int) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    dates = cast(
        pd.DatetimeIndex,
        pd.to_datetime([row[0] for row in rows[data_start_row - 1 :]], errors="coerce"),
    )
    valid_dates = cast(pd.DatetimeIndex, dates[~pd.isna(dates)])
    date_start = cast(pd.Timestamp, valid_dates.min())
    date_end = cast(pd.Timestamp, valid_dates.max())
    return {
        "name": sheet_name,
        "rows": len(rows),
        "columns": len(rows[0]) if rows else 0,
        "data_rows": len(valid_dates),
        "date_start": date_start.date().isoformat(),
        "date_end": date_end.date().isoformat(),
        "duplicate_dates": int(valid_dates.duplicated().sum()),
        "monotonic_dates": bool(valid_dates.is_monotonic_increasing),
        "headers": list(rows[1]) if len(rows) > 1 else [],
    }


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Profile Raw_wind.xlsx")
    parser.add_argument("--workbook", default=None)
    parser.add_argument("--run-id", default=None)
    args = parser.parse_args(argv)
    config = load_config()
    bundle = load_manual_excel_bundle(config, args.workbook)
    run_id = args.run_id or f"manual_{bundle.workbook_sha256[:10]}"
    report_dir = Path("reports") / "wind_macro_daily" / "data_audits" / run_id
    report_dir.mkdir(parents=True, exist_ok=True)

    market_index = cast(pd.DatetimeIndex, bundle.market_data.index)
    market_start = cast(pd.Timestamp, market_index.min())
    market_end = cast(pd.Timestamp, market_index.max())
    macro_series = cast(pd.Series, bundle.macro_observations["series_id"])
    profile = {
        "workbook_filename": bundle.workbook_path.name,
        "workbook_sha256": bundle.workbook_sha256,
        "sheets": [
            _sheet_profile(bundle.workbook_path, "Price_raw", 4),
            _sheet_profile(bundle.workbook_path, "Macro", 5),
        ],
        "active_universe": bundle.active_universe,
        "market_panel": {
            "rows": len(bundle.market_data),
            "start": market_start.date().isoformat(),
            "end": market_end.date().isoformat(),
            "columns": list(bundle.market_data.columns),
        },
        "macro": {
            "usable_series": int(macro_series.nunique()),
            "observations": len(bundle.macro_observations),
            "strict_point_in_time_eligible": False,
            "reason": "current workbook contains no historical vintages",
        },
    }
    (report_dir / "workbook_profile.json").write_text(
        json.dumps(profile, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    _write_csv(bundle.price_quality, report_dir / "price_quality.csv", index=False)
    _write_csv(bundle.macro_update_audit, report_dir / "macro_update_audit.csv", index=False)
    gaps = pd.DataFrame(
        [
            {
                "required_item": "US_IRS_5Y",
                "status": "missing",
                "needed_data": "5Y SOFR OIS/IRS par rate, not Treasury yield",
                "frequency": "daily close",
            },
            {
                "required_item": "CN_IRS_5Y",
                "status": "missing",
                "needed_data": "5Y FR007 IRS par rate",
                "frequency": "daily close",
            },
            {
                "required_item": "FX carry inputs",
                "status": "missing",
                "needed_data": "matched-tenor USD/CNY/JPY/AUD/EUR OIS or forward points",
                "frequency": "daily close",
            },
            {
                "required_item": "historical macro vintages",
                "status": "missing",
                "needed_data": "reference date, exact release timestamp, vintage id, value",
                "frequency": "each release and revision",
            },
        ]
    )
    _write_csv(gaps, report_dir / "required_data_gaps.csv", index=False)
    print(f"wrote workbook audit to {report_dir}")


if __name__ == "__main__":
    main()
